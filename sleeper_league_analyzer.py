#!/usr/bin/env python3
"""
Sleeper Fantasy Football League Analyzer
Queries a Sleeper league and outputs teams sorted by projected points
with starter status information.
"""

import requests
import json
import csv
from typing import Dict, List, Any
import sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

class SleeperLeagueAnalyzer:
    def __init__(self, league_id: str):
        self.league_id = league_id
        self.base_url = "https://api.sleeper.app/v1"
        
    def get_league_info(self) -> Dict[str, Any]:
        """Get basic league information"""
        url = f"{self.base_url}/league/{self.league_id}"
        print(f"API Request: {url}")
        response = requests.get(url)
        response.raise_for_status()
        return response.json()
    
    def get_users(self) -> List[Dict[str, Any]]:
        """Get all users in the league"""
        url = f"{self.base_url}/league/{self.league_id}/users"
        print(f"API Request: {url}")
        response = requests.get(url)
        response.raise_for_status()
        return response.json()
    
    def get_rosters(self) -> List[Dict[str, Any]]:
        """Get all rosters in the league"""
        url = f"{self.base_url}/league/{self.league_id}/rosters"
        print(f"API Request: {url}")
        response = requests.get(url)
        response.raise_for_status()
        return response.json()
    
    def get_current_week(self) -> int:
        """Get current NFL week"""
        url = f"{self.base_url}/state/nfl"
        response = requests.get(url)
        response.raise_for_status()
        return response.json()["week"]
    
    def get_matchups(self, week: int) -> List[Dict[str, Any]]:
        """Get matchups for a specific week"""
        url = f"{self.base_url}/league/{self.league_id}/matchups/{week}"
        print(f"API Request: {url}")
        response = requests.get(url)
        response.raise_for_status()
        return response.json()
    
    def get_players(self) -> Dict[str, Any]:
        """Get all NFL players data"""
        url = f"{self.base_url}/players/nfl"
        response = requests.get(url)
        response.raise_for_status()
        return response.json()
    
    def get_nfl_state(self) -> Dict[str, Any]:
        """Get current NFL state including week and season info"""
        url = f"{self.base_url}/state/nfl"
        print(f"API Request: {url}")
        response = requests.get(url)
        response.raise_for_status()
        return response.json()
    
    def get_weekly_projections(self, week: int) -> Dict[str, Any]:
        """
        Get weekly projections using the correct Sleeper API format
        """
        # Use the correct sleeper.com endpoint with parameters
        url = f"https://api.sleeper.com/projections/nfl/2025/{week}"
        params = {
            'season_type': 'regular',
            'position[]': ['DEF', 'FLEX', 'QB', 'RB', 'SUPER_FLEX', 'TE', 'WR'],
            'order_by': 'ppr'
        }
        
        try:
            print(f"API Request: {url} with params: {params}")
            response = requests.get(url, params=params)
            if response.status_code == 200:
                data = response.json()
                print(f"✓ Found projections: {len(data)} players")
                
                # Convert list format to dict format keyed by player_id for easier lookup
                projections_dict = {}
                for player_data in data:
                    player_id = player_data.get('player_id')
                    if player_id:
                        projections_dict[player_id] = player_data
                
                # Show sample data - find players with actual projections
                players_with_projections = []
                for player_data in data[:50]:  # Check first 50 players
                    stats = player_data.get('stats', {})
                    if any(key in stats for key in ['pts_ppr', 'fantasy_points_ppr', 'projected_points', 'pts_std']):
                        players_with_projections.append(player_data)
                        if len(players_with_projections) >= 3:
                            break
                
                if players_with_projections:
                    print("Players with projection data:")
                    for player_data in players_with_projections:
                        player_name = f"{player_data.get('player', {}).get('first_name', '')} {player_data.get('player', {}).get('last_name', '')}"
                        stats = player_data.get('stats', {})
                        print(f"  {player_name}: {stats}")
                else:
                    print("No players found with projection stats in first 50 entries")
                
                return projections_dict
            else:
                print(f"✗ No data at: {url} (Status: {response.status_code})")
        except Exception as e:
            print(f"✗ Error at: {url} - {e}")
        
        return {}
    
    def get_nfl_schedule(self, week: int) -> Dict[str, Any]:
        """Get NFL schedule for a specific week"""
        url = f"{self.base_url}/state/nfl"
        print(f"API Request: {url}")
        response = requests.get(url)
        response.raise_for_status()
        return response.json()
    
    def check_player_game_status(self, player_id: str, week: int, players_data: Dict[str, Any]) -> bool:
        """
        Check if a player's game has started/finished for the week
        This checks if the player has any points scored (indicating game started)
        """
        # Simple heuristic: if player has points in current matchup, their game has started
        # This is not perfect but gives us a basic indication
        
        # We'll check if the player has any points scored this week
        # If they have points > 0, we assume their game has started
        
        # This is a simplified approach that works for live scoring
        # A more sophisticated approach would check actual NFL game times
        return False  # Will be updated with matchup data

def main():
    if len(sys.argv) != 2:
        print("Usage: python sleeper_league_analyzer.py <league_id>")
        sys.exit(1)
    
    league_id = sys.argv[1]
    analyzer = SleeperLeagueAnalyzer(league_id)
    
    try:
        # Get league data
        print("Fetching league data...")
        league_info = analyzer.get_league_info()
        users = analyzer.get_users()
        rosters = analyzer.get_rosters()
        nfl_state = analyzer.get_nfl_state()
        current_week = nfl_state['week']
        matchups = analyzer.get_matchups(current_week)
        players_data = analyzer.get_players()
        
        # Try to find weekly projections
        print(f"\nSearching for weekly projections for week {current_week}...")
        weekly_projections = analyzer.get_weekly_projections(current_week)
        
        if weekly_projections:
            print(f"Weekly projections found for {len(weekly_projections)} players")
            
            # Find players with actual projection data
            players_with_data = []
            for player_id, proj_data in weekly_projections.items():
                if proj_data and len(proj_data) > 0:
                    players_with_data.append((player_id, proj_data))
                if len(players_with_data) >= 3:
                    break
            
            if players_with_data:
                print("Sample projection data:")
                for player_id, proj_data in players_with_data:
                    print(f"  Player {player_id}: {proj_data}")
            else:
                print("No players found with projection data")
        else:
            print("No weekly projections found - using season totals")
        
        print(f"League: {league_info['name']}")
        print(f"Current Week: {current_week}")
        print(f"Teams: {len(rosters)}")
        
        # Debug: Show what's available in roster data
        if rosters:
            print(f"\nDebug - Available roster data keys: {list(rosters[0].keys())}")
            # Show a sample roster with some data
            sample_roster = rosters[0]
            print(f"Sample roster data:")
            for key, value in sample_roster.items():
                if key == 'players' and isinstance(value, list):
                    print(f"  {key}: [{len(value)} players]")
                elif key == 'starters' and isinstance(value, list):
                    print(f"  {key}: [{len(value)} starters]")
                else:
                    print(f"  {key}: {value}")
        
        # Create user lookup
        user_lookup = {user['user_id']: user for user in users}
        
        # Process team data
        teams_data = []
        for roster in rosters:
            owner_id = roster.get('owner_id')
            user = user_lookup.get(owner_id, {})
            team_name = user.get('display_name', f"Team {roster['roster_id']}")
            
            # Find matchup data for this roster
            matchup_data = next((m for m in matchups if m['roster_id'] == roster['roster_id']), {})
            
            # Get current points
            current_points = matchup_data.get('points', 0)
            starters_points = matchup_data.get('starters_points', [])
            total_current = sum(starters_points) if starters_points else current_points
            
            # Count starters and check game status
            starters = roster.get('starters', [])
            starters_played = 0
            
            # Calculate weekly projected points from individual player projections
            weekly_projected = 0
            
            if weekly_projections:
                for player_id in starters:
                    if player_id and player_id in weekly_projections:
                        player_proj_data = weekly_projections[player_id]
                        stats = player_proj_data.get('stats', {})
                        
                        # Look for various projection fields
                        player_proj = (stats.get('pts_ppr', 0) or 
                                     stats.get('fantasy_points_ppr', 0) or
                                     stats.get('projected_points', 0) or
                                     stats.get('pts_std', 0))
                        
                        weekly_projected += player_proj
            
            # Use weekly projections if available, otherwise fall back to current points
            projected_points = weekly_projected if weekly_projected > 0 else total_current
            
            # Check each starter to see if their game has started
            # Use matchup data to see if players have scored points
            players_points = matchup_data.get('players_points', {})
            
            for player_id in starters:
                if player_id:
                    # If player has scored any points, assume their game has started
                    player_points = players_points.get(player_id, 0)
                    if player_points > 0:
                        starters_played += 1
            
            teams_data.append({
                'team_name': team_name,
                'roster_id': roster['roster_id'],
                'current_points': total_current,
                'projected_points': projected_points,
                'starters_total': len([s for s in starters if s]),  # Count non-null starters
                'starters_played': starters_played
            })
        
        # Filter out teams with 0 projections (eliminated teams)
        active_teams = [team for team in teams_data if team['projected_points'] > 0]
        
        # Sort by projected points ascending (lowest first - most at risk)
        active_teams.sort(key=lambda x: x['projected_points'])
        
        # Output as both CSV and styled Excel
        csv_file = f"sleeper_league_{league_id}_week_{current_week}.csv"
        excel_file = f"sleeper_league_{league_id}_week_{current_week}.xlsx"
        
        # Create CSV file
        with open(csv_file, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['Risk Rank', 'Team Name', 'Current Points', 'Projected Points', 'Starters Played', 'Total Starters']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            writer.writeheader()
            for i, team in enumerate(active_teams, 1):
                writer.writerow({
                    'Risk Rank': i,
                    'Team Name': team['team_name'],
                    'Current Points': team['current_points'],
                    'Projected Points': team['projected_points'],
                    'Starters Played': team['starters_played'],
                    'Total Starters': team['starters_total']
                })
        
        # Create styled Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = f"Week {current_week} Risk Analysis"
        
        # Define colors
        red_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")  # Light red
        yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")  # Light yellow
        green_fill = PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid")  # Light green
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light gray
        bold_font = Font(bold=True)
        
        # Headers
        headers = ['Risk Rank', 'Team Name', 'Current Points', 'Projected Points', 'Starters Played', 'Total Starters']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = bold_font
        
        # Data rows with color coding
        for i, team in enumerate(active_teams, 1):
            row = i + 1  # +1 because row 1 is headers
            
            # Determine fill color based on risk rank
            if i <= 2:  # Top 2 most at risk - red
                fill_color = red_fill
            elif i <= 5:  # Next 3 - yellow
                fill_color = yellow_fill
            else:  # Rest - green
                fill_color = green_fill
            
            # Add data to cells
            data = [
                i,  # Risk Rank
                team['team_name'],
                team['current_points'],
                round(team['projected_points'], 2),
                team['starters_played'],
                team['starters_total']
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.fill = fill_color
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(excel_file)
        
        print(f"\nResults saved to:")
        print(f"  CSV: {csv_file}")
        print(f"  Excel (styled): {excel_file}")
        
        # Also print to console (only active teams)
        print(f"\nActive Teams Risk Rankings (Lowest Projected Points First):")
        print(f"Active Teams: {len(active_teams)} | Eliminated: {len(teams_data) - len(active_teams)}")
        print("-" * 80)
        for i, team in enumerate(active_teams, 1):
            print(f"{i:2d}. {team['team_name']:<20} | Current: {team['current_points']:6.1f} | Proj: {team['projected_points']:6.1f} | {team['starters_played']}/{team['starters_total']} played")
            
    except requests.exceptions.RequestException as e:
        print(f"Error fetching data from Sleeper API: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()